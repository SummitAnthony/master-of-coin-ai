"""Enhanced Excel advisory workbook (openpyxl).

Numbers come only from ``Facts`` (already quantized); prose is rendered verbatim
from the narrative. ``statement`` is optional and used only to enrich detailed
opex line-item audit rows.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from advisor import PRODUCT_NAME
from advisor.narrative.advisor import Advisory
from advisor.schema import EntityMeta, Facts, IncomeStatement, PeriodKPIs

from .formatting import BRAND_PRIMARY_HEX, DISCLAIMER, excel_number_format, status_hex

_HEADER_FILL = PatternFill("solid", fgColor=BRAND_PRIMARY_HEX)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=16, color=BRAND_PRIMARY_HEX)

# (row label — "{unit}" is replaced with the entity's volume unit,
#  PeriodKPIs attribute, number-format kind)
_PNL_ROWS: tuple[tuple[str, str, str], ...] = (
    ("Revenue", "revenue", "money"),
    ("Cost of Goods Sold", "cogs", "money"),
    ("Gross Profit", "gross_profit", "money"),
    ("Total Operating Expenses", "total_opex", "money"),
    ("Operating Profit", "operating_profit", "money"),
    ("Other Income", "other_income", "money"),
    ("Finance Cost", "finance_cost", "money"),
    ("Profit Before Tax", "profit_before_tax", "money"),
    ("Tax Expense", "tax_expense", "money"),
    ("Net Profit", "net_profit", "money"),
    ("Sales Volume", "volume_mt", "volume"),
)
_KPI_ROWS: tuple[tuple[str, str, str], ...] = (
    ("Gross margin", "gross_margin_pct", "pct"),
    ("Operating margin", "operating_margin_pct", "pct"),
    ("Net margin", "net_margin_pct", "pct"),
    ("COGS to sales", "cogs_to_sales_pct", "pct"),
    ("Opex to sales", "opex_to_sales_pct", "pct"),
    ("Finance cost to sales", "finance_cost_to_sales_pct", "pct"),
    ("Selling price / {unit}", "selling_price_per_mt", "per_unit"),
    ("COGS / {unit}", "cogs_per_mt", "per_unit"),
    ("Gross profit / {unit}", "gross_profit_per_mt", "per_unit"),
)


def _f(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _header_row(ws: Worksheet, labels: list[str], start_col: int = 2) -> None:
    cell = ws.cell(row=1, column=1, value="Metric")
    cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
    for i, label in enumerate(labels):
        c = ws.cell(row=1, column=start_col + i, value=label)
        c.fill, c.font = _HEADER_FILL, _HEADER_FONT


def _matrix_sheet(
    ws: Worksheet,
    kpis: list[PeriodKPIs],
    rows: tuple[tuple[str, str, str], ...],
    entity: EntityMeta,
) -> None:
    _header_row(ws, [k.period.label for k in kpis])
    for r, (label, attr, kind) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=label.replace("{unit}", entity.volume_unit))
        for c, k in enumerate(kpis, start=2):
            cell = ws.cell(row=r, column=c, value=_f(getattr(k, attr)))
            cell.number_format = excel_number_format(
                kind,  # type: ignore[arg-type]
                currency=entity.currency,
                unit=entity.volume_unit,
            )


def _title_line(entity: EntityMeta) -> str:
    if entity.group_name:
        return f"{entity.group_name} — {entity.company_name}"
    return entity.company_name


def _cover(ws: Worksheet, facts: Facts, narrative: Advisory, generated_at: datetime) -> None:
    ws["A1"] = _title_line(facts.entity)
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"{PRODUCT_NAME} — Financial Advisory"
    ws["A3"] = f"Generated: {generated_at.isoformat()}"
    ws["A4"] = f"Latest period: {facts.latest_period.label}"
    ws["A5"] = f"Narrative: {narrative.provider}/{narrative.model}" + (
        " (degraded)" if narrative.degraded else ""
    )
    ws["A7"] = "Executive Summary"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = narrative.executive_summary
    ws["A10"] = DISCLAIMER


def _status_sheet(ws: Worksheet, facts: Facts) -> None:
    _header_row(ws, ["Period", "Value", "Status", "Code"])
    ws.cell(row=1, column=1, value="Metric").fill = _HEADER_FILL
    for r, s in enumerate(facts.statuses, start=2):
        ws.cell(row=r, column=1, value=s.metric)
        ws.cell(row=r, column=2, value=s.period)
        ws.cell(row=r, column=3, value=_f(s.value))
        status_cell = ws.cell(row=r, column=4, value=s.status.value)
        status_cell.fill = PatternFill("solid", fgColor=status_hex(s.status))
        status_cell.font = Font(color="FFFFFF", bold=True)
        ws.cell(row=r, column=5, value=s.message_code)


def _variance_sheet(ws: Worksheet, facts: Facts) -> None:
    headers = ["Metric", "Basis", "From", "To", "Abs change", "Pct change", "Bps", "Direction"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
    for r, v in enumerate(facts.variances, start=2):
        ws.cell(row=r, column=1, value=v.metric)
        ws.cell(row=r, column=2, value=v.basis.value)
        ws.cell(row=r, column=3, value=v.from_period)
        ws.cell(row=r, column=4, value=v.to_period)
        ws.cell(row=r, column=5, value=_f(v.absolute_change))
        ws.cell(row=r, column=6, value=_f(v.pct_change))
        ws.cell(row=r, column=7, value=_f(v.bps_change))
        ws.cell(row=r, column=8, value=v.direction.value)


def _anomaly_sheet(ws: Worksheet, facts: Facts) -> None:
    headers = ["Code", "Severity", "Period", "Metric", "Observed", "Threshold", "Message"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
    for r, a in enumerate(facts.anomalies, start=2):
        ws.cell(row=r, column=1, value=a.code)
        ws.cell(row=r, column=2, value=a.severity.value)
        ws.cell(row=r, column=3, value=a.period)
        ws.cell(row=r, column=4, value=a.metric)
        ws.cell(row=r, column=5, value=_f(a.observed))
        ws.cell(row=r, column=6, value=_f(a.threshold))
        ws.cell(row=r, column=7, value=a.message_code)


def _add_margin_chart(ws: Worksheet, n_periods: int) -> None:
    chart = LineChart()
    chart.title = "Margin trend"
    chart.y_axis.title = "Percent"
    # First three KPI rows are the margins; columns 2..n+1 are periods.
    data = Reference(ws, min_col=1, max_col=1 + n_periods, min_row=2, max_row=4)
    chart.add_data(data, from_rows=True, titles_from_data=True)
    cats = Reference(ws, min_col=2, max_col=1 + n_periods, min_row=1, max_row=1)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{len(_KPI_ROWS) + 4}")


def build_workbook(
    facts: Facts,
    narrative: Advisory,
    *,
    generated_at: datetime,
    statement: IncomeStatement | None = None,
) -> Workbook:
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    _cover(cover, facts, narrative, generated_at)
    _matrix_sheet(wb.create_sheet("Income Statement"), facts.kpis, _PNL_ROWS, facts.entity)
    kpi_ws = wb.create_sheet("KPIs")
    _matrix_sheet(kpi_ws, facts.kpis, _KPI_ROWS, facts.entity)
    if len(facts.kpis) >= 2:
        _add_margin_chart(kpi_ws, len(facts.kpis))
    _variance_sheet(wb.create_sheet("Variance"), facts)
    _status_sheet(wb.create_sheet("Status"), facts)
    _anomaly_sheet(wb.create_sheet("Anomalies"), facts)
    return wb


def write_excel(
    facts: Facts,
    narrative: Advisory,
    path: Path,
    *,
    generated_at: datetime,
    statement: IncomeStatement | None = None,
) -> Path:
    build_workbook(facts, narrative, generated_at=generated_at, statement=statement).save(path)
    return path
