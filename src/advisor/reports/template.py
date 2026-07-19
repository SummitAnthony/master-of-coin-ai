"""Blank .xlsx input template showing exactly where the data goes.

The generated workbook round-trips through ``extract_income_statement``: the
row labels are canonical aliases the label-matcher recognises, the period
headers parse as fiscal years, and the example figures are internally
consistent. An Instructions sheet documents the accepted labels, period
formats, and scale markers.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from advisor.ingestion.labels import ALIASES
from advisor.schema import EntityMeta

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_NOTE_FONT = Font(italic=True, color="808080")

# (row label, first-period example, second-period example). Gross profit is
# revenue - cogs exactly, so the template passes the schema consistency check.
_EXAMPLE_ROWS: tuple[tuple[str, int, int], ...] = (
    ("Revenue", 1_800_000, 2_100_000),
    ("Cost of Goods Sold", 1_170_000, 1_390_000),
    ("Gross Profit", 630_000, 710_000),
    ("Selling and Distribution Expenses", 120_000, 140_000),
    ("Administrative Expenses", 95_000, 105_000),
    ("Other Operating Expenses", 25_000, 30_000),
    ("Operating Profit", 390_000, 435_000),
    ("Other Income", 12_000, 15_000),
    ("Finance Cost", 48_000, 61_000),
    ("Profit Before Tax", 354_000, 389_000),
    ("Tax Expense", 88_500, 97_250),
    ("Net Profit", 265_500, 291_750),
    ("Sales Volume", 1_500, 1_680),
)

_REQUIRED_LABELS = ("Revenue", "Cost of Goods Sold")

_INSTRUCTIONS: tuple[str, ...] = (
    "How to use this template",
    "",
    "1. Put one reporting period per column on the 'Income Statement' sheet, with the",
    "   period name in the header row. Accepted header formats: FY2024, FY2023-24,",
    "   Q1 FY2024, H1 FY2024, Jan 2024, 2024.",
    "2. Put one line item per row, with its label in column A. Only Revenue and",
    "   Cost of Goods Sold are required; every other row is optional.",
    "3. Replace the example figures with your own. Enter absolute amounts, or add a",
    '   note such as "in \'000" above the header row to have all figures scaled.',
    "4. Sales Volume is entered in your configured volume unit and is never scaled.",
    "5. Rows with unrecognised labels are preserved as extra lines (not discarded).",
    "",
    "Recognised labels per line (case-insensitive):",
)


def _alias_lines() -> list[str]:
    return [f"  {key.value}: {', '.join(aliases)}" for key, aliases in ALIASES.items()]


def _statement_sheet(ws: Worksheet, entity: EntityMeta) -> None:
    ws["A1"] = f"{entity.company_name} — Income Statement"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Amounts in {entity.currency} (absolute); volume in {entity.volume_unit}"
    ws["A2"].font = _NOTE_FONT

    header_row = 4
    for col, text in enumerate(("Particulars", "FY2024", "FY2025"), start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
    for offset, (label, first, second) in enumerate(_EXAMPLE_ROWS):
        r = header_row + 1 + offset
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=first)
        ws.cell(row=r, column=3, value=second)
        if label in _REQUIRED_LABELS:
            ws.cell(row=r, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14


def _instructions_sheet(ws: Worksheet) -> None:
    for r, line in enumerate((*_INSTRUCTIONS, *_alias_lines()), start=1):
        ws.cell(row=r, column=1, value=line)
    ws.cell(row=1, column=1).font = _TITLE_FONT
    ws.column_dimensions["A"].width = 100


def build_input_template(entity: EntityMeta | None = None) -> Workbook:
    """Build the blank template workbook (example figures included)."""
    entity = entity if entity is not None else EntityMeta()
    wb = Workbook()
    statement = wb.active
    assert statement is not None  # a fresh Workbook always has an active sheet
    statement.title = "Income Statement"
    _statement_sheet(statement, entity)
    _instructions_sheet(wb.create_sheet("Instructions"))
    return wb


def write_input_template(path: Path, entity: EntityMeta | None = None) -> Path:
    """Write the template workbook to ``path`` and return it."""
    build_input_template(entity).save(path)
    return path
