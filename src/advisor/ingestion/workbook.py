"""The ONLY module that imports openpyxl.

Loads an .xlsx into an immutable in-memory :class:`SheetGrid` so all downstream
extraction logic is pure and unit-testable without binary fixtures. File handles
are always closed (Windows file-locking safe).
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from .errors import SheetNotFoundError, WorkbookReadError


@dataclass(frozen=True)
class Cell:
    """A single spreadsheet cell with 0-based row/col and an A1-style coord."""

    row: int
    col: int
    coord: str
    value: object


@dataclass(frozen=True)
class SheetGrid:
    """An immutable grid of cells. ``rows[i][j]`` is the cell at row i, col j."""

    title: str
    rows: tuple[tuple[Cell, ...], ...]


def grid_from_rows(title: str, raw_rows: list[list[object]]) -> SheetGrid:
    """Build a SheetGrid from a plain list-of-lists (test helper, no openpyxl)."""
    rows = tuple(
        tuple(
            Cell(row=r, col=c, coord=f"{get_column_letter(c + 1)}{r + 1}", value=value)
            for c, value in enumerate(raw_row)
        )
        for r, raw_row in enumerate(raw_rows)
    )
    return SheetGrid(title=title, rows=rows)


def load_grid(path: Path, *, sheet: str | None = None) -> SheetGrid:
    """Read an .xlsx file into a :class:`SheetGrid`, closing the workbook."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, OSError, zipfile.BadZipFile) as exc:
        raise WorkbookReadError(path, str(exc)) from exc
    try:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise SheetNotFoundError(sheet, list(wb.sheetnames))
            ws = wb[sheet]
        else:
            ws = wb.active
        title = ws.title
        # Compute indices via enumerate: read-only blank cells are EmptyCell
        # objects that lack .row/.column/.coordinate attributes.
        rows = tuple(
            tuple(
                Cell(
                    row=r,
                    col=c,
                    coord=f"{get_column_letter(c + 1)}{r + 1}",
                    value=cell.value,
                )
                for c, cell in enumerate(row)
            )
            for r, row in enumerate(ws.iter_rows())
        )
    finally:
        wb.close()
    return SheetGrid(title=title, rows=rows)
